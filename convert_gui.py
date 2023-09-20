import re
import tkinter as tk
from tkinter import filedialog
import fitz
import openpyxl as px
from openpyxl.styles import Alignment
import os


def modify_text(text):
    # 不要な改行の削除
    mod_text = text.replace("\n", "")
    # 文末の句点の後に改行を追加する
    mod_text = re.sub(r"([。])([」』]?)", r"\1\2\n", mod_text)
    # 文字コードにない文字(文字化け等)の削除
    mod_text = mod_text.encode("cp932", errors="ignore").decode("cp932")

    return mod_text


# PDFファイルを変換する関数
def convert_pdf_to_excel():
    label_text.set("変換中")
    # PDFファイルを選択
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    # PDFテキストを格納するリスト
    txt_list = []
    
    # result_label.config(text="変換中")
   
    if file_path:
        # PDFファイルを読み込む
        doc = fitz.open(file_path)

        # PDFテキストを取得し、リストに格納
        for page in range(len(doc)):
            text = doc[page].get_text()
            text = modify_text(text)
            txt_list.append([page + 1, text])

        # 新しいExcelファイルを作成
        wb = px.Workbook()
        ws = wb.active

        # Excelの書式設定
        my_alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions["B"].width = 100

        # Excelのヘッダーを出力
        headers = ["ページ", "内容"]
        for i, header in enumerate(headers):
            ws.cell(row=1, column=1 + i, value=headers[i])

        # ExcelにPDFのテキストを出力
        for y, row in enumerate(txt_list):
            for x, cell in enumerate(row):
                ws.cell(row=y + 2, column=x + 1, value=cell)
                ws.cell(row=y + 2, column=x + 1).alignment = my_alignment

        # Excelファイルを保存
        script_dir = "./"  # スクリプトのディレクトリ取得
        output_dir = os.path.join(script_dir, "converted")  # 出力先ディレクトリパス作成
        os.makedirs(output_dir, exist_ok=True)  # ディレクトリが存在しない場合は作成
        excel_file_path = os.path.join(
            output_dir, os.path.basename(file_path).replace(".pdf", "_excel.xlsx")
        )
        wb.save(excel_file_path)

        # 完了メッセージを表示
        label_text.set(f"変換完了: {excel_file_path}")
        # result_label.config(text=f"変換完了: {excel_file_path}")


# Tkinterウィンドウを作成し、サイズを設定
root = tk.Tk()
root.title("PDF to Excel Converter")
root.geometry("300x150")  # ウィンドウのサイズを指定

# 変換ボタンを作成
convert_button = tk.Button(root, text="PDFをExcelに変換", command=convert_pdf_to_excel)
convert_button.pack(pady=20)

# 結果を表示するラベル
label_text = tk.StringVar(root)
label_text.set("")
result_label = tk.Label(root, textvariable=label_text, wraplength=400)
result_label.pack()


# アプリケーションのメインループ
root.mainloop()
