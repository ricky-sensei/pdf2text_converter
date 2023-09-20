# １：ライブラリ設定
import fitz  # pymupdfライブラリ
import openpyxl as px
from openpyxl.styles import Alignment

# ２：PDFテキストを格納するリスト作成
txt_list = []

# ３：PDFファイルを読み込む
filename = "ninen_jyo.pdf"
doc = fitz.open("./pdf/" + filename)

# ４：１ページずつテキストデータを取得
for page in range(len(doc)):
    text = doc[page].get_text()
    text = text.replace("\n", "")
    txt_list.append([page + 1, text])

# ５：新しいExcelファイルを作成
wb = px.Workbook()
ws = wb.active

# ６：Excelの書式設定
my_alignment = Alignment(vertical="top", wrap_text=True)
ws.column_dimensions["B"].width = 100

# ７：Excelのヘッダーを出力
headers = ["ページ", "内容"]
for i, header in enumerate(headers):
    ws.cell(row=1, column=1 + i, value=headers[i])

# ８：ExcelにPDFのテキストを出力
for y, row in enumerate(txt_list):
    for x, cell in enumerate(row):
        ws.cell(row=y + 2, column=x + 1, value=cell)
        ws.cell(row=y + 2, column=x + 1).alignment = my_alignment

# ９：Excelファイルの保存
xlname = "./converted/" + filename.split(".")[0] + "_excel.xlsx"
wb.save(xlname)
